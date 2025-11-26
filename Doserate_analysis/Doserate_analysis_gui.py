# TempCorrection Program
# Park Kilsoon

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib
matplotlib.rc('font', family='Malgun Gothic')  # 윈도우 기본 한글 폰트
matplotlib.rcParams['axes.unicode_minus'] = False  # 마이너스(-) 깨짐 방지
from matplotlib.widgets import Slider, Button
from tkinter import Tk, filedialog, messagebox, Menu
import os
import sys
import io
import tkinter as tk
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

if sys.stdout and hasattr(sys.stdout, "detach"):
    sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding='utf-8')

# 전역 변수
file_path = None
df = None
g_values = None
selected_points = []
window_size = 30

class GMApp:
    def __init__(self, root):
        self.root = root
        self.root.title("GM/PIN CPS 분석기")
        self.file_path = None
        self.df = None
        self.g_values = None
        self.selected_points = []
        self.window_size = 30
        self.is_log_scale = tk.BooleanVar(value=True)

        # 메뉴바
        menubar = Menu(root)
        filemenu = Menu(menubar, tearoff=0)
        filemenu.add_command(label="파일 열기", command=self.open_file)
        menubar.add_cascade(label="파일", menu=filemenu)
        root.config(menu=menubar)

        # 상단 프레임(툴바+저장버튼)
        top_frame = tk.Frame(root)
        top_frame.pack(side=tk.TOP, fill=tk.X)

        # matplotlib Figure
        self.fig, (self.ax, self.ax_info) = plt.subplots(1, 2, figsize=(12, 6), gridspec_kw={'width_ratios': [4, 1]})
        self.canvas = FigureCanvasTkAgg(self.fig, master=root)
        self.canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)

        # 툴바(아이콘) - top_frame에 배치
        self.toolbar = NavigationToolbar2Tk(self.canvas, top_frame)
        self.toolbar.update()
        self.toolbar.pack(side=tk.LEFT, padx=2, pady=2)
        # 툴바의 좌표 표시 비활성화
        self.toolbar.set_message = lambda s: None

        # 저장 버튼 - top_frame에 배치
        self.btn_save = tk.Button(top_frame, text="저장", command=self.save_selected_points)
        self.btn_save.pack(side=tk.LEFT, padx=10, pady=5)

        # 로그 스케일 체크박스
        self.log_scale_check = tk.Checkbutton(top_frame, text="로그 스케일", variable=self.is_log_scale, command=self.redraw_plot)
        self.log_scale_check.pack(side=tk.LEFT, padx=5)

        # x, y 좌표 표시 라벨 (저장 버튼 오른쪽)
        self.xy_label = tk.Label(top_frame, text="", width=24, anchor=tk.W)
        self.xy_label.pack(side=tk.LEFT, padx=10)

        # 슬라이더 (Tkinter 위젯)
        control_frame = tk.Frame(root)
        control_frame.pack(side=tk.BOTTOM, fill=tk.X)
        tk.Label(control_frame, text="윈도우 크기 (±)").pack(side=tk.LEFT)
        self.scale = tk.Scale(control_frame, from_=10, to=100, orient=tk.HORIZONTAL, command=self.update_slider)
        self.scale.set(30)
        self.scale.pack(side=tk.LEFT, fill=tk.X, expand=1)

        # 상태바 (진행상태 표시)
        self.status_var = tk.StringVar()
        self.status_var.set("")
        self.status_bar = tk.Label(root, textvariable=self.status_var, bd=1, relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

        self.canvas.mpl_connect('button_press_event', self.on_click)
        self.fig.canvas.mpl_connect('motion_notify_event', self.on_mouse_move)

        # 안내 메시지
        print("상단 메뉴에서 '파일 > 파일 열기'를 선택하세요.")

        # __init__ 마지막에 추가
        self.ax_info.axis('off')
        self.canvas.draw()

        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def open_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not path:
            return
        self.file_path = path
        try:
            self.g_values = self.read_g_column_with_progress(self.file_path)
            if self.g_values is None or len(self.g_values) == 0:
                self.selected_points.clear()
                self.ax.clear()
                self.ax_info.clear()
                self.ax_info.axis('off')
                self.canvas.draw()
                self.status_var.set("G열 데이터가 없습니다.")
                messagebox.showwarning("경고", "G열 데이터가 없습니다.")
                return
            self.selected_points.clear()
            self.redraw_plot()
        except Exception as e:
            self.status_var.set("")
            messagebox.showerror("오류", f"엑셀 파일을 읽는 중 오류 발생:\n{e}")

    def read_g_column_with_progress(self, file_path):
        wb = load_workbook(file_path, read_only=True)
        ws = wb['Sheet1']
        total_rows = ws.max_row - 1  # 헤더 제외
        g_values = []
        for i, row in enumerate(ws.iter_rows(min_row=2, min_col=7, max_col=7, values_only=True), 1):
            val = row[0]
            # 값이 없으면 0, 음수면 None(제거)
            if val is None or (isinstance(val, float) and np.isnan(val)):
                g_values.append(0)
            elif isinstance(val, (int, float)) and val < 0:
                g_values.append(0)  # 음수도 0으로
            else:
                g_values.append(val)
            if i % 100 == 0 or i == total_rows:
                percent = int(i / total_rows * 100)
                self.status_var.set(f"엑셀 파일 읽는 중... ({percent}%)")
                self.root.update_idletasks()
        wb.close()
        self.status_var.set("엑셀 파일 읽기 완료.")
        # 인덱스를 엑셀 행 번호(2번 행부터)로 맞추고 싶으면:
        s = pd.Series(g_values, index=range(2, 2 + len(g_values)))
        return s

    def on_click(self, event):
        # 툴바의 확대/이동 기능 활성화 시 인덱스 선택 방지
        if self.toolbar.mode != '':
            return

        if self.g_values is None or event.xdata is None:
            return

        clicked_index = int(round(event.xdata))

        # 선택된 윈도우 영역 내부를 클릭하면 해당 인덱스 선택 해제
        for selected_idx in self.selected_points[:]:
            if (selected_idx - self.window_size) <= clicked_index <= (selected_idx + self.window_size):
                self.selected_points.remove(selected_idx)
                print(f"인덱스 {selected_idx} 선택 해제됨.")
                self.redraw_plot()
                return

        # 새 인덱스 선택 (경계값 확인)
        first_valid_index = self.g_values.index[0] + self.window_size
        last_valid_index = self.g_values.index[-1] - self.window_size
        if not (first_valid_index <= clicked_index <= last_valid_index):
            print(f"선택한 인덱스({clicked_index})가 너무 끝쪽입니다.")
            return
        
        self.selected_points.append(clicked_index)
        print(f"인덱스 {clicked_index} 선택됨.")
        self.redraw_plot()

    def redraw_plot(self):
        self.ax.clear()
        self.ax_info.clear()
        if self.g_values is not None:
            self.ax.plot(self.g_values.index, self.g_values, marker='o', linestyle='None', markersize=2, label='G 값')
            
            # 파일명(확장자 제외)로 그래프 제목 설정
            if self.file_path:
                base_title = os.path.splitext(os.path.basename(self.file_path))[0]
            else:
                base_title = ''
            if self.is_log_scale.get():
                self.ax.set_yscale('log')
                self.ax.set_title(f'{base_title} (로그 스케일)')
            else:
                self.ax.set_yscale('linear')
                self.ax.set_title(f'{base_title} (리니어 스케일)')

            self.ax.set_xlabel('Index')
            self.ax.set_ylabel('G 값 (avgGM)')
            for idx in self.selected_points:
                self.ax.axvline(idx, color='red', linestyle='--', linewidth=0.8)
                # 녹색 투명 사각형 박스 추가
                left = idx - self.window_size
                right = idx + self.window_size
                # 인덱스 범위 내로 제한
                left = max(left, self.g_values.index[0])
                right = min(right, self.g_values.index[-1])
                self.ax.add_patch(
                    plt.Rectangle(
                        (left, self.ax.get_ylim()[0]),
                        right - left,
                        self.ax.get_ylim()[1] - self.ax.get_ylim()[0],
                        linewidth=2,
                        edgecolor='red',
                        facecolor='green',
                        alpha=0.2,
                        zorder=1
                    )
                )
            # 표 형태로 표시 (항상 인덱스 오름차순 정렬, No 추가)
            if self.selected_points:
                sorted_points = sorted(self.selected_points)
                table_data = [[i+1, idx, f"{self.g_values[idx-self.window_size:idx+self.window_size+1].mean():.8f}"] for i, idx in enumerate(sorted_points)]
                self.ax_info.axis('off')
                table = self.ax_info.table(
                    cellText=table_data,
                    colLabels=["No", "Index", "MeanValue"],
                    loc='center',
                    cellLoc='center'
                )
                table.auto_set_font_size(False)
                table.set_fontsize(12)
                table.auto_set_column_width(col=list(range(len(table_data[0]))))  # 열 너비 자동 조정
                # 각 열의 width와 각 행의 height를 1.2배로 늘림
                for k, cell in table.get_celld().items():
                    cell.set_width(cell.get_width() * 1.2)
                    cell.set_height(cell.get_height() * 1.2)
            else:
                self.ax_info.axis('off')
                self.ax_info.text(0.5, 0.5, "선택된 인덱스 없음", va='center', ha='center', fontsize=12)
        self.canvas.draw()

    def save_selected_points(self):
        if not self.selected_points:
            messagebox.showwarning("저장 실패", "선택된 인덱스가 없습니다.")
            return

        results = []
        # 항상 인덱스 오름차순으로 저장
        for idx in sorted(self.selected_points):
            window = self.g_values[idx - self.window_size:idx + self.window_size + 1]
            mean_val = window.mean()
            results.append((idx, mean_val))

        df_result = pd.DataFrame(results, columns=["Index", "MeanValue"])
        df_result.insert(0, "No", range(1, len(df_result) + 1))

        # 새로운 파일명 생성
        base = os.path.splitext(os.path.basename(self.file_path))[0]
        dir_path = os.path.dirname(self.file_path)
        new_file = os.path.join(dir_path, base + "_CPS.xlsx")
        image_path = os.path.join(dir_path, base + ".png")

        try:
            # 1. 그래프 이미지를 먼저 저장
            self.save_plot_image()
            # 2. 표 저장
            df_result.to_excel(new_file, sheet_name='평균결과', index=False)
            # 3. 이미지 삽입
            wb = load_workbook(new_file)
            ws = wb['평균결과']
            img_row = len(df_result) + 3  # 표 아래 2줄 띄우고
            img = XLImage(image_path)
            # 이미지 크기 지정 (단위: cm)
            img.width = 24 * 37.7952755906  # 1cm = 37.795... pixel
            img.height = 12 * 37.7952755906
            ws.add_image(img, f'A{img_row}')
            wb.save(new_file)
            print(f"평균결과와 그래프 이미지가 {new_file} 파일에 저장되었습니다.")
            messagebox.showinfo("저장 완료", f"평균결과와 그래프 이미지가 {new_file} 파일에 저장되었습니다.")
        except Exception as e:
            messagebox.showerror("저장 오류", f"엑셀 저장 중 오류:\n{e}")

    def save_plot_image(self):
        # 엑셀 파일명과 동일하게, 확장자만 png로 변경
        base = os.path.splitext(os.path.basename(self.file_path))[0]
        dir_path = os.path.dirname(self.file_path)
        image_path = os.path.join(dir_path, base + ".png")
        self.fig.savefig(image_path, dpi=300)
        print(f"그래프 이미지 저장 완료: {image_path}")

    def update_slider(self, val):
        self.window_size = int(val)
        self.redraw_plot()

    def on_mouse_move(self, event):
        # x, y 좌표 표시를 저장 버튼 오른쪽 라벨에 표시
        if event.inaxes == self.ax:
            if event.xdata is not None and event.ydata is not None:
                self.xy_label.config(text=f"x={event.xdata:.0f}, y={event.ydata:.4f}")
            else:
                self.xy_label.config(text="")
        else:
            self.xy_label.config(text="")

    def on_close(self):
        # 모든 matplotlib Figure 닫기
        plt.close('all')
        # Tkinter 창 닫기
        self.root.destroy()
        # 프로세스 강제 종료
        os._exit(0)

if __name__ == "__main__":
    root = tk.Tk()
    app = GMApp(root)
    root.mainloop()
