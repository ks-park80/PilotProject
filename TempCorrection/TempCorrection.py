# TempCorrection Program
# Park Kilsoon

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import time
import io

import numpy as np
import pandas as pd
import matplotlib
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from matplotlib import font_manager as fm
import openpyxl
from openpyxl.drawing.image import Image as XLImage

from scipy.signal import savgol_filter
from scipy.ndimage import gaussian_filter1d
from sklearn.metrics import r2_score

def set_korean_font():
    candidates = ["Malgun Gothic", "맑은 고딕", "NanumGothic", "AppleGothic"]
    for f in candidates:
        try:
            if fm.findfont(f, fallback_to_default=False):
                matplotlib.rcParams["font.family"] = f
                break
        except:
            continue
    matplotlib.rcParams["axes.unicode_minus"] = False

set_korean_font()

class SpectrumAnalyzer:
    def __init__(self, root):
        self.root = root
        self.root.title("스펙트럼 분석기")
        self.root.geometry("1500x900")

        self.temp_targets = [-20, -10, 0, 10, 20, 30, 40, 50]
        self.temp_data = None
        self.spectrum_all = None
        self.selected_points = {}      # key=x_idx
        self.checkbox_vars = {}
        self.roi_center = None
        self.roi_minus = tk.IntVar(value=70)
        self.roi_plus = tk.IntVar(value=70)
        self.poly_degree = tk.IntVar(value=4)
        self.ratio_var = tk.DoubleVar(value=1.0)
        self.chpeak_b_var = tk.DoubleVar(value=1.0)
        self.ref_val = 0.0
        self.last_equation = ""
        self.last_r2 = None

        # 그래프 x축 입력값
        self.top_xmin_var = tk.StringVar()
        self.top_xmax_var = tk.StringVar()
        self.bottom_xmin_var = tk.StringVar()
        self.bottom_xmax_var = tk.StringVar()

        # 마우스 커서 위치값 레이블 변수(상단/하단)
        self.top_cursor_var = tk.StringVar()
        self.bottom_cursor_var = tk.StringVar()

        # 마우스 커서 추적용 세로선 객체들
        self.cursor_line_top = None    # 상단 그래프 커서선
        self.cursor_line_bottom = None # 하단 그래프 커서선

        self._build_ui()

    def _build_ui(self):
        top = tk.Frame(self.root)
        top.pack(fill="x", pady=5)
        tk.Button(top, text="엑셀 파일 열기", command=self.open_excel).pack(side="left", padx=5)
        tk.Button(top, text="결과 저장", command=self.save_excel).pack(side="left", padx=5)
        tk.Button(top, text="선택 초기화", command=self.reset_selections).pack(side="left", padx=5)

        tk.Label(top, text="Ratio :").pack(side="left")
        tk.Entry(top, width=15, textvariable=self.ratio_var).pack(side="left", padx=(0, 10))
        tk.Label(top, text="chpeak_b :").pack(side="left")
        tk.Entry(top, width=15, textvariable=self.chpeak_b_var).pack(side="left", padx=(0, 10))
        tk.Button(top, text="채널 에러 계산", command=self.calculate_channel_error).pack(side="left", padx=5)
        self.ref_val_label = tk.Label(top, text="ref_val_CH: 0")
        self.ref_val_label.pack(side="left", padx=(5, 15))
        tk.Label(top, text="다항식 차수:").pack(side="left")
        ttk.Combobox(top, textvariable=self.poly_degree, values=[2,3,4,5,6], width=3, state="readonly").pack(side="left")
        tk.Button(top, text="수식 계산", command=self.calculate_polynomial).pack(side="left", padx=5)
        self.progress_var = tk.StringVar()
        tk.Label(top, textvariable=self.progress_var, fg="blue").pack(side="left", padx=15)

        roi_frame = tk.LabelFrame(top, text="ROI 범위(채널)")
        roi_frame.pack(side="right", padx=10)
        tk.Label(roi_frame, text="–").pack(side="left")
        tk.Spinbox(roi_frame, from_=0, to=1024, width=5, textvariable=self.roi_minus, command=self._update_peak_channels).pack(side="left", padx=2)
        tk.Label(roi_frame, text="+").pack(side="left")
        tk.Spinbox(roi_frame, from_=0, to=1024, width=5, textvariable=self.roi_plus, command=self._update_peak_channels).pack(side="left")

        # ⭐ 한 줄로 배치되는 Min/Max/적용/커서값
        xscale_line = tk.Frame(self.root)
        xscale_line.pack(fill="x", padx=5, pady=(0,3))
        tk.Label(xscale_line, text="상단 X축(Interval): Min").pack(side="left")
        tk.Entry(xscale_line, width=8, textvariable=self.top_xmin_var).pack(side="left")
        tk.Label(xscale_line, text="Max").pack(side="left")
        tk.Entry(xscale_line, width=8, textvariable=self.top_xmax_var).pack(side="left")
        tk.Button(xscale_line, text="적용", command=self._plot_temp).pack(side="left", padx=(2,6))
        tk.Label(xscale_line, text="| ").pack(side="left")
        tk.Label(xscale_line, text="하단 X축(채널): Min").pack(side="left")
        tk.Entry(xscale_line, width=8, textvariable=self.bottom_xmin_var).pack(side="left")
        tk.Label(xscale_line, text="Max").pack(side="left")
        tk.Entry(xscale_line, width=8, textvariable=self.bottom_xmax_var).pack(side="left")
        tk.Button(xscale_line, text="적용", command=self._plot_spectrum).pack(side="left", padx=(2,6))
        tk.Label(xscale_line, text="| 현재 위치:").pack(side="left")
        self.cursor_pos_label = tk.Label(xscale_line, textvariable=self.top_cursor_var)
        self.cursor_pos_label.pack(side="left", padx=5)
        tk.Label(xscale_line, text=" / ").pack(side="left")
        self.bottom_cursor_label = tk.Label(xscale_line, textvariable=self.bottom_cursor_var)
        self.bottom_cursor_label.pack(side="left")

        main = tk.Frame(self.root)
        main.pack(fill="both", expand=True, padx=10, pady=10)
        left = tk.Frame(main)
        left.pack(side="left", fill="both", expand=True)
        self.fig = Figure(figsize=(12,8), dpi=100)
        self.ax1 = self.fig.add_subplot(211)
        self.ax2 = self.fig.add_subplot(212)
        self.fig.subplots_adjust(hspace=0.35)
        self._init_axes()
        self.canvas = FigureCanvasTkAgg(self.fig, master=left)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(fill="both", expand=True)
        # 마우스 위치 표시 이벤트 연결
        self.canvas.mpl_connect("motion_notify_event", self._on_mouse_move)
        self.canvas.mpl_connect("button_press_event", self.on_click)

        right = tk.Frame(main)
        right.pack(side="right", fill="y")
        table_frame = tk.Frame(right)
        table_frame.pack(fill="both", expand=False)
        tk.Label(table_frame, text="선택 지점", font=("Arial",12,"bold")).pack(pady=5)
        cols = ("No", "Temp(°C)", "Interval", "Temp_ADC", "Peak_CH", "Err_CH", "Spc", "표시")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=10)
        for c in cols:
            self.tree.heading(c, text=c)
            width = 60 if c == "No" else (80 if c not in ("표시", "Temp(°C)") else 55)
            self.tree.column(c, anchor="center", width=width, stretch=False)
        self.tree.pack(side="left", fill="both", expand=True)
        ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview).pack(side="right", fill="y")
        self.tree.bind("<Double-1>", self.toggle_checkbox)
        graph_frame = tk.Frame(right)
        graph_frame.pack(fill="both", expand=True, pady=(10,0))
        self.fig3 = Figure(figsize=(5,3), dpi=100)
        self.ax3 = self.fig3.add_subplot(111)
        self.canvas3 = FigureCanvasTkAgg(self.fig3, master=graph_frame)
        self.canvas3.get_tk_widget().pack(fill="both", expand=True)

    def _on_mouse_move(self, event):
        """마우스 이동 시 커서 위치 표시 + 빨간 세로선 표시"""
        if event.inaxes == self.ax1 and self.temp_data is not None and event.xdata is not None:
            idx = int(round(event.xdata))
            if 0 <= idx < len(self.temp_data):
                y = self.temp_data[idx]
                self.top_cursor_var.set(f"x={idx:.0f} / y={y:.2f}")
                
                # 상단 그래프에 빨간 세로선 표시
                if self.cursor_line_top is not None:
                    self.cursor_line_top.remove()
                self.cursor_line_top = self.ax1.axvline(x=event.xdata, color='red', linestyle='-', alpha=0.7, linewidth=1)
                self.canvas.draw_idle()  # 성능을 위해 draw_idle 사용
            else:
                self.top_cursor_var.set("")
                if self.cursor_line_top is not None:
                    self.cursor_line_top.remove()
                    self.cursor_line_top = None
                    self.canvas.draw_idle()
                    
        elif event.inaxes == self.ax2 and self.spectrum_all is not None and event.xdata is not None:
            idx = int(round(event.xdata))
            if 0 <= idx < self.spectrum_all.shape[1]:
                # 여러 선택지점 중 첫번째의 스펙트럼(있는 경우만)
                if self.selected_points:
                    any_idx = next(iter(self.selected_points))
                    y = self.selected_points[any_idx]["spectrum"][idx]
                    self.bottom_cursor_var.set(f"x={idx:.0f} / y={y:.2f}")
                else:
                    self.bottom_cursor_var.set(f"x={idx:.0f}")
                
                # 하단 그래프에 빨간 세로선 표시
                if self.cursor_line_bottom is not None:
                    self.cursor_line_bottom.remove()
                self.cursor_line_bottom = self.ax2.axvline(x=event.xdata, color='red', linestyle='-', alpha=0.7, linewidth=1)
                self.canvas.draw_idle()
            else:
                self.bottom_cursor_var.set("")
                if self.cursor_line_bottom is not None:
                    self.cursor_line_bottom.remove()
                    self.cursor_line_bottom = None
                    self.canvas.draw_idle()
        else:
            # 그래프 밖으로 마우스가 나간 경우
            self.top_cursor_var.set("")
            self.bottom_cursor_var.set("")
            
            # 두 그래프의 커서선 모두 제거
            if self.cursor_line_top is not None:
                self.cursor_line_top.remove()
                self.cursor_line_top = None
            if self.cursor_line_bottom is not None:
                self.cursor_line_bottom.remove()
                self.cursor_line_bottom = None
            self.canvas.draw_idle()

    def _init_axes(self):
        self.ax1.clear()
        self.ax1.set_title("온도 데이터")
        self.ax1.set_xlabel("Interval Time")
        self.ax1.set_ylabel("Temp. ADC")
        self.ax1.grid(True)
        self.ax2.clear()
        self.ax2.set_title("스펙트럼 데이터 (2048채널)")
        self.ax2.set_xlabel("채널")
        self.ax2.set_ylabel("값")
        self.ax2.grid(True)

    def open_excel(self):
        path = filedialog.askopenfilename(
            title="엑셀 파일 선택", filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not path:
            return
        threading.Thread(target=self._load_excel, args=(path,), daemon=True).start()

    def _update_progress(self, pct):
        self.progress_var.set(f"엑셀파일을 읽는중... {pct}%")
        self.root.update_idletasks()

    def _load_excel(self, path):
        try:
            self._update_progress(0)
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            n_rows = ws.max_row
            data = []
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                data.append(row[:2058])
                if i % 100 == 0 or i == n_rows:
                    self._update_progress(int(i / n_rows * 90))
            wb.close()
            df = pd.DataFrame(data)
            self._update_progress(95)
            self.temp_data = df.iloc[:, 8].to_numpy()
            self.spectrum_all = df.iloc[:, 10:2058].to_numpy()
            self._plot_temp()
            self._update_progress(100)
            time.sleep(0.4)
            self.progress_var.set("")
        except Exception as e:
            self.progress_var.set("")
            messagebox.showerror("오류", f"파일 읽기 오류:\n{e}")

    def _plot_temp(self, redraw_temp_data=True):
        """온도 그래프 그리기 (커서선 유지 로직 추가)"""
        try:
            xxmin = float(self.top_xmin_var.get())
        except:
            xxmin = 0
        try:
            xxmax = float(self.top_xmax_var.get())
        except:
            xxmax = len(self.temp_data)-1 if self.temp_data is not None else 0
            
        if redraw_temp_data:
            # 커서선 초기화
            self.cursor_line_top = None
            
            self.ax1.clear()
            self.ax1.plot(np.arange(len(self.temp_data)), self.temp_data, "b-", linewidth=1)
            self.ax1.set_title("온도 데이터")
            self.ax1.set_xlabel("Interval Time")
            self.ax1.set_ylabel("Temp. ADC")
            self.ax1.grid(True)
        else:
            lines_to_remove = []
            for line in self.ax1.lines:
                xdata = line.get_xdata()
                # 선택지점 수직선(빨간 점선)과 커서선(빨간 실선) 구분
                if (len(xdata) == 2 and xdata[0] == xdata[1] and 
                    line.get_color() == 'red' and line.get_linestyle() == '--'):
                    lines_to_remove.append(line)
            for line in lines_to_remove:
                line.remove()
        
        # 선택된 지점의 수직선 다시 그리기 (점선으로)
        for x_idx in sorted(self.selected_points.keys()):
            self.ax1.axvline(x_idx, color="red", linestyle="--")
        
        self.ax1.set_xlim(xxmin, xxmax)
        self.canvas.draw()

    def on_click(self, event):
        if event.inaxes == self.ax1 and self.temp_data is not None and event.xdata is not None:
            self._handle_top_click(int(round(event.xdata)))
        elif event.inaxes == self.ax2 and self.spectrum_all is not None and event.xdata is not None:
            self._handle_bottom_click(int(round(event.xdata)))

    def _handle_top_click(self, x_idx):
        if not (0 <= x_idx < len(self.temp_data)):
            return
        for existing in list(self.selected_points.keys()):
            if abs(existing - x_idx) <= 10:
                del self.selected_points[existing]
                del self.checkbox_vars[existing]
                self.tree.delete(str(existing))
                self._refresh_table()
                self._plot_temp()
                self._plot_spectrum()
                return
        if len(self.selected_points) >= 8:
            messagebox.showinfo("안내", "최대 8개 지점 선택 가능합니다.")
            return
        y = self.temp_data[x_idx]
        nearest_temp = self.temp_targets[len(self.selected_points)]
        raw_spec = self.spectrum_all[x_idx, :]
        sg_spec = savgol_filter(raw_spec, window_length=11, polyorder=3)
        filtered_spec = gaussian_filter1d(sg_spec, sigma=2)
        self.selected_points[x_idx] = {
            "temp": nearest_temp,
            "adc": y,
            "spectrum": filtered_spec,
            "peak_ch": None,
            "spc": None
        }
        self.ax1.axvline(x_idx, color="red", linestyle="--")
        self.canvas.draw()
        self._insert_row(x_idx, self.selected_points[x_idx])
        self._refresh_table()
        self._plot_temp()
        self._plot_spectrum()

    def _insert_row(self, x_idx, info):
        self.selected_points = dict(sorted(self.selected_points.items(), key=lambda kv: kv[0]))
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.checkbox_vars.clear()
        for i, (idx, val) in enumerate(self.selected_points.items(), start=1):
            temp_fixed = self.temp_targets[i-1]
            var = tk.BooleanVar(value=True)
            self.checkbox_vars[idx] = var
            var.trace_add("write", lambda *_: self._plot_spectrum())
            self.tree.insert(
                "", "end", iid=str(idx),
                values=(
                    i,
                    f"{temp_fixed}",
                    idx,
                    f"{val['adc']:.2f}",
                    val.get("peak_ch", ""),
                    val.get("err_ch", ""),
                    val.get("spc", ""),
                    "☑"
                )
            )

    def _refresh_table(self):
        self.selected_points = dict(sorted(self.selected_points.items(), key=lambda kv: kv[0]))
        for item in self.tree.get_children():
            self.tree.delete(item)
        for i, (idx, val) in enumerate(self.selected_points.items(), start=1):
            temp_fixed = self.temp_targets[i-1]
            var = self.checkbox_vars.get(idx, tk.BooleanVar(value=True))
            self.tree.insert(
                "", "end", iid=str(idx),
                values=(
                    i,
                    f"{temp_fixed}",
                    idx,
                    f"{val['adc']:.2f}",
                    val.get("peak_ch", ""),
                    val.get("err_ch", ""),
                    val.get("spc", ""),
                    "☑" if var.get() else "☐"
                )
            )

    def toggle_checkbox(self, _):
        sel = self.tree.focus()
        if not sel:
            return
        idx = int(sel)
        var = self.checkbox_vars.get(idx)
        if not var:
            return
        var.set(not var.get())
        self.tree.set(sel, "표시", "☑" if var.get() else "☐")
        self._refresh_table()
        self._plot_temp()
        self._plot_spectrum()

    def _handle_bottom_click(self, ch_idx):
        self.roi_center = ch_idx
        self._plot_spectrum()
        self._update_peak_channels()

    def _plot_spectrum(self):
        """스펙트럼 그래프 그리기 (커서선 유지 로직 추가)"""
        # 커서선 초기화
        self.cursor_line_bottom = None
        
        self.ax2.clear()
        colors = plt.cm.tab10.colors
        
        try:
            xxmin = float(self.bottom_xmin_var.get())
        except:
            xxmin = 0
        try:
            xxmax = float(self.bottom_xmax_var.get())
        except:
            xxmax = self.spectrum_all.shape[1]-1 if self.spectrum_all is not None else 2047
            
        if self.roi_center is not None:
            left = max(self.roi_center - self.roi_minus.get(), 0)
            right = min(self.roi_center + self.roi_plus.get(), self.spectrum_all.shape[1] - 1)
            self.ax2.axvspan(left, right, color='gray', alpha=0.3)
            self.ax2.axvline(self.roi_center, color="red", linestyle="--")
            
        for i, (x_idx, info) in enumerate(self.selected_points.items()):
            if self.checkbox_vars[x_idx].get():
                spec = info["spectrum"]
                self.ax2.plot(
                    np.arange(len(spec)), spec,
                    color=colors[i % len(colors)],
                    label=f"{info['temp']}°C",
                    linewidth=1
                )
                
        self.ax2.set_title("스펙트럼 데이터 (2048채널)")
        self.ax2.set_xlabel("채널")
        self.ax2.set_ylabel("값")
        self.ax2.grid(True)
        self.ax2.set_xlim(xxmin, xxmax)
        
        if self.ax2.lines:
            self.ax2.legend()
        self.canvas.draw()

    def _update_peak_channels(self, *_):
        if self.roi_center is None:
            return
        m, p = self.roi_minus.get(), self.roi_plus.get()
        for idx, info in self.selected_points.items():
            spec = info["spectrum"]
            l = max(self.roi_center - m, 0)
            r = min(self.roi_center + p, len(spec) - 1)
            roi = spec[l:r+1]
            off = int(np.argmax(roi))
            peak = l + off
            spc_val = int(spec[peak])
            info["peak_ch"], info["spc"] = peak, spc_val
            self.tree.set(str(idx), "Peak_CH", peak)
            self.tree.set(str(idx), "Spc", spc_val)

    def reset_selections(self):
        """선택 초기화 (커서선도 함께 초기화)"""
        if not self.selected_points:
            messagebox.showinfo("안내", "현재 선택된 지점이 없습니다.")
            return
            
        self.selected_points.clear()
        self.checkbox_vars.clear()
        
        # 커서선 초기화
        self.cursor_line_top = None
        self.cursor_line_bottom = None
        
        # x축 입력 초기화
        self.top_xmin_var.set("")
        self.top_xmax_var.set("")
        self.bottom_xmin_var.set("")
        self.bottom_xmax_var.set("")
        
        # 커서 위치 표시 초기화
        self.top_cursor_var.set("")
        self.bottom_cursor_var.set("")
        
        self._plot_temp(redraw_temp_data=False)
        
        self.ax2.clear()
        self.ax2.set_title("스펙트럼 데이터 (2048채널)")
        self.ax2.set_xlabel("채널")
        self.ax2.set_ylabel("값")
        self.ax2.grid(True)
        self.ax2.set_xlim(0, self.spectrum_all.shape[1]-1 if self.spectrum_all is not None else 2047)
        self.canvas.draw()
        
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        self.ax3.clear()
        self.ax3.set_title("Temp_ADC vs Peak_CH 회귀 분석")
        self.ax3.set_xlabel("Temp_ADC")
        self.ax3.set_ylabel("Peak_CH")
        self.ax3.grid(True)
        self.canvas3.draw()
        
        self.last_equation = ""
        self.last_r2 = None 

    def calculate_channel_error(self):
        if not self.selected_points:
            messagebox.showinfo("안내", "계산할 데이터가 없습니다.")
            return
        try:
            ref_val = float(self.ratio_var.get()) * float(self.chpeak_b_var.get())
        except ValueError:
            messagebox.showerror("오류", "Ratio 또는 chpeak_b 값이 유효하지 않습니다.")
            return
        self.ref_val_label.config(text=f"ref_val_CH: {int(ref_val)}")
        for idx, info in self.selected_points.items():
            peak = info.get("peak_ch")
            if peak is None:
                continue
            err_rate = (peak - ref_val) / ref_val * 100
            info["err_ch"] = f"{err_rate:.2f}"
            self.tree.set(str(idx), "Err_CH", info["err_ch"])
        messagebox.showinfo("완료", "채널 오차 계산이 완료되었습니다.")

    def calculate_polynomial(self):
        x=[]; y=[]
        for idx, info in self.selected_points.items():
            if self.checkbox_vars[idx].get() and info.get("peak_ch") is not None:
                x.append(info["adc"]); y.append(info["peak_ch"])
        deg = self.poly_degree.get()
        if len(x) < deg+1:
            messagebox.showerror("오류", "데이터가 부족합니다.")
            return
        x_arr=np.array(x); y_arr=np.array(y)
        coeffs = np.polyfit(x_arr, y_arr, deg)
        p = np.poly1d(coeffs)
        y_pred = p(x_arr)
        r2 = r2_score(y_arr, y_pred)
        terms=[]
        for i,c in enumerate(coeffs):
            power=deg-i
            terms.append(f"{c:.20f}x^{power}")
        equation=" + ".join(terms)
        self.last_equation = equation
        self.last_r2 = r2
        lines = []
        for i in range(0, len(terms), 2):
            pair = " + ".join(terms[i:i+2])
            lines.append(pair)
        equation_graph = " +\n  ".join(lines)
        self.ax3.clear()
        self.ax3.scatter(x_arr, y_arr, color="blue", label="데이터")
        xs=np.linspace(x_arr.min(), x_arr.max(), 200)
        self.ax3.plot(xs, p(xs), 'r--', label=f"{deg}차 회귀선")
        self.ax3.set_title("Temp_ADC vs Peak_CH 회귀 분석")
        self.ax3.set_xlabel("Temp_ADC"); self.ax3.set_ylabel("Peak_CH"); self.ax3.legend()
        # 수식/R2를 그래프의 ‘아래쪽’에 표시
        text=f"y = {equation_graph}\nR² = {r2:.4f}"
        self.ax3.text(0.15,0.02,text,transform=self.ax3.transAxes,
                      fontsize=8, va="bottom",
                      bbox=dict(boxstyle="round",facecolor="white",alpha=0.7))
        self.canvas3.draw()

    def save_excel(self):
        if not self.selected_points:
            messagebox.showwarning("경고", "저장할 데이터가 없습니다.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        out = []
        ordered = list(self.selected_points.items())
        for i, (idx, info) in enumerate(ordered, start=1):
            if self.checkbox_vars[idx].get():
                out.append({
                    "No": i,
                    "Temperature": info["temp"],
                    "Interval_Time": idx,
                    "Temp_ADC": info["adc"],
                    "Peak_CH": info["peak_ch"],
                    "Err_CH": info.get("err_ch", ""),
                    "Spc": info["spc"],
                    "Spectrum": info["spectrum"].tolist()
                })
        df_out = pd.DataFrame(out)
        df_out.to_excel(path, index=False)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        ref_val = float(self.ratio_var.get()) * float(self.chpeak_b_var.get())
        start_row = len(df_out) + 2
        ws.cell(row=start_row, column=1, value="Ratio")
        ws.cell(row=start_row, column=2, value=self.ratio_var.get())
        ws.cell(row=start_row, column=4, value="chpeak_b")
        ws.cell(row=start_row, column=5, value=self.chpeak_b_var.get())
        ws.cell(row=start_row, column=7, value="ref_val")
        ws.cell(row=start_row, column=8, value=ref_val)
        # 수식·R2 저장
        eq = getattr(self, 'last_equation', '')
        r2v= getattr(self, 'last_r2', None)
        start_row = len(df_out) + 4
        ws.cell(row=start_row, column=1, value="y : ")
        ws.cell(row=start_row, column=2, value=self.last_equation)
        ws.cell(row=start_row+1, column=1, value="R2 : ")
        ws.cell(row=start_row+1, column=2, value=self.last_r2)
        # 그래프 이미지 저장
        buf1 = io.BytesIO()
        self.ax1.figure.savefig(buf1, format='png')
        buf1.seek(0)
        img1 = XLImage(buf1)
        buf2 = io.BytesIO()
        self.ax2.figure.savefig(buf2, format='png')
        buf2.seek(0)
        img2 = XLImage(buf2)
        buf3 = io.BytesIO()
        self.ax3.figure.savefig(buf3, format='png')
        buf3.seek(0)
        img3 = XLImage(buf3)
        start_row = len(df_out) + 7
        img1.anchor = f'A{start_row}'
        img3.anchor = f'A{start_row + 40}'
        ws.add_image(img1)
        ws.add_image(img3)
        wb.save(path)
        messagebox.showinfo("완료", "결과와 그래프 이미지가 저장되었습니다.")

def main():
    root = tk.Tk()
    SpectrumAnalyzer(root)
    root.mainloop()

if __name__ == "__main__":
    main()
